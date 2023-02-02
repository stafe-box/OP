from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class calculator:
    def calculate(pwrs, index, time, cost):
        index = str(index)
        return round(pwrs[index]*time*cost / 1000.00, 2)


    def calc_to_text(pwrs, index, time, cost):
        dic = {
            '0': 'Утюг',
            '1': 'Телевизор',
            '2': 'Стиральная машина'
        }
        index = str(index)
        return (dic[index]+' потребляет '+str(calculator.calculate(pwrs, index, time, cost))+' рыбов')

class docx_saver:
    
    def save_to_docx(text_result):
        file_name = './result.docx'
        document = Document(file_name)
        document.add_paragraph(text_result)
        document.save(file_name)

class xlsx_saver:
    def save_to_xlsx(name, result):
        if result > 0:
            dic = {
                '0': 'Утюг',
                '1': 'Телевизор',
                '2': 'Стиральная машина'
            }
            name = dic[str(name)]
            wb = load_workbook('./result.xlsx')
            ws = wb.active
            i = 1
            for row in ws.iter_rows(values_only=True):  # geting count of rows
                i += 1
            print('i:', i)
            ws['A'+str(i)] = name
            ws['B'+str(i)] = result
            wb.save('./result.xlsx')

class Tecnic:
    name = 'Техника'
    def __init__(self, pwr):
        self.pwr = pwr

    def calculate(self, time, cost):
        return round(self.pwr*time*cost / 1000.00, 2)

    def calculate_to_text(self, time, cost):
        return (f"{self.name} потребляет {self.calculate(time, cost)} рыбов")

class iron(Tecnic):
    name = 'Утюг'
    def _init_(self, pwr):
        super().__init__(pwr)

    def calculate(self, time, cost):
        return super().calculate(time, cost)

    def calculate_to_text(self, time, cost):
        return super().calculate_to_text(time, cost)


class TV(Tecnic):
    name = 'Телевизор'
    def _init_(self, pwr):
        super().__init__(pwr)

    def calculate(self, time, cost):
        return super().calculate(time, cost)

    def calculate_to_text(self, time, cost):
        return super().calculate_to_text(time, cost)

class washer(Tecnic):
    name = 'Утюг'
    def _init_(self, pwr):
        super().__init__(pwr)

    def calculate(self, time, cost):
        return super().calculate(time, cost)

    def calculate_to_text(self, time, cost):
        return super().calculate_to_text(time, cost)