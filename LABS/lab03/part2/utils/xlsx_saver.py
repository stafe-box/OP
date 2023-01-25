from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def save_to_xlsx(name, result):
    if result > 0:
        dic = {
            '0': 'Утюг',
            '1': 'Телевизор',
            '2': 'Стиральная машина'
        }
        name = dic[str(name)]
        wb = load_workbook('./result.xlsx')
        ws = wb.active
        i = 1
        for row in ws.iter_rows(values_only=True):  # geting count of rows
            i += 1
        print('i:', i)
        ws['A'+str(i)] = name
        ws['B'+str(i)] = result
        wb.save('./result.xlsx')
